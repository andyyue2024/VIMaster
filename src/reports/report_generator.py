"""
报告生成模块 - 支持 PDF、Excel 格式和自定义模板
"""
import logging
import os
from typing import Dict, Any, List, Optional
from dataclasses import dataclass, field
from datetime import datetime
from abc import ABC, abstractmethod
import json

logger = logging.getLogger(__name__)

# 检查可选依赖
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("reportlab 不可用，PDF 生成功能将被禁用。请运行: pip install reportlab")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl 不可用，Excel 生成功能将被禁用。请运行: pip install openpyxl")


@dataclass
class ReportTemplate:
    """报告模板配置"""
    name: str = "default"
    title: str = "VIMaster 价值投资分析报告"
    subtitle: str = ""
    author: str = "VIMaster 系统"
    logo_path: Optional[str] = None

    # 页面设置
    page_size: str = "A4"  # A4 或 letter
    margin_top: float = 2.0  # cm
    margin_bottom: float = 2.0
    margin_left: float = 2.5
    margin_right: float = 2.0

    # 样式设置
    primary_color: str = "#1a5f7a"
    secondary_color: str = "#57c5b6"
    text_color: str = "#333333"

    # 内容设置
    include_summary: bool = True
    include_financials: bool = True
    include_valuation: bool = True
    include_moat: bool = True
    include_risk: bool = True
    include_signals: bool = True
    include_ml_score: bool = True
    include_charts: bool = False  # 图表需要额外依赖

    # 页脚
    footer_text: str = "本报告由 VIMaster 自动生成，仅供参考"
    show_page_numbers: bool = True
    show_generation_time: bool = True

    def to_dict(self) -> Dict[str, Any]:
        return {
            "name": self.name,
            "title": self.title,
            "subtitle": self.subtitle,
            "author": self.author,
            "page_size": self.page_size,
            "primary_color": self.primary_color,
            "include_summary": self.include_summary,
            "include_financials": self.include_financials,
            "include_valuation": self.include_valuation,
            "include_moat": self.include_moat,
            "include_risk": self.include_risk,
            "include_signals": self.include_signals,
            "include_ml_score": self.include_ml_score,
        }

    @staticmethod
    def from_dict(data: Dict[str, Any]) -> "ReportTemplate":
        return ReportTemplate(**{k: v for k, v in data.items() if hasattr(ReportTemplate, k)})

    def save(self, path: str) -> None:
        os.makedirs(os.path.dirname(path) if os.path.dirname(path) else ".", exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self.to_dict(), f, ensure_ascii=False, indent=2)

    @staticmethod
    def load(path: str) -> "ReportTemplate":
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return ReportTemplate.from_dict(data)


@dataclass
class StockReportData:
    """单只股票报告数据"""
    stock_code: str
    stock_name: str = ""

    # 财务指标
    current_price: Optional[float] = None
    pe_ratio: Optional[float] = None
    pb_ratio: Optional[float] = None
    roe: Optional[float] = None
    gross_margin: Optional[float] = None
    debt_ratio: Optional[float] = None
    free_cash_flow: Optional[float] = None

    # 估值
    intrinsic_value: Optional[float] = None
    fair_price: Optional[float] = None
    margin_of_safety: Optional[float] = None
    valuation_score: Optional[float] = None

    # 护城河
    moat_score: Optional[float] = None
    brand_strength: Optional[float] = None
    cost_advantage: Optional[float] = None

    # 风险
    risk_level: str = ""
    leverage_risk: Optional[float] = None

    # 信号
    buy_signal: str = ""
    sell_signal: str = ""
    final_signal: str = ""

    # 评分
    overall_score: Optional[float] = None
    ml_score: Optional[float] = None

    # 决策
    decision: str = ""
    position_size: Optional[float] = None
    stop_loss: Optional[float] = None
    take_profit: Optional[float] = None


@dataclass
class PortfolioReportData:
    """投资组合报告数据"""
    report_id: str = ""
    generated_at: str = ""
    stocks: List[StockReportData] = field(default_factory=list)

    # 统计
    total_stocks: int = 0
    strong_buy_count: int = 0
    buy_count: int = 0
    hold_count: int = 0
    sell_count: int = 0
    strong_sell_count: int = 0

    # 组合建议
    strategy: str = ""
    expected_return: Optional[float] = None
    risk_score: Optional[float] = None


class BaseReportGenerator(ABC):
    """报告生成器基类"""

    def __init__(self, template: Optional[ReportTemplate] = None):
        self.template = template or ReportTemplate()

    @abstractmethod
    def generate_single_stock_report(self, data: StockReportData, output_path: str) -> bool:
        """生成单只股票报告"""
        pass

    @abstractmethod
    def generate_portfolio_report(self, data: PortfolioReportData, output_path: str) -> bool:
        """生成投资组合报告"""
        pass


class PDFReportGenerator(BaseReportGenerator):
    """PDF 报告生成器"""

    def __init__(self, template: Optional[ReportTemplate] = None):
        super().__init__(template)
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab 不可用，请运行: pip install reportlab")

    def _get_page_size(self):
        return A4 if self.template.page_size == "A4" else letter

    def _get_styles(self):
        styles = getSampleStyleSheet()

        # 自定义样式
        styles.add(ParagraphStyle(
            name='ChineseTitle',
            fontSize=24,
            leading=30,
            alignment=1,  # 居中
            spaceAfter=20,
            textColor=colors.HexColor(self.template.primary_color),
        ))

        styles.add(ParagraphStyle(
            name='ChineseSubtitle',
            fontSize=14,
            leading=18,
            alignment=1,
            spaceAfter=30,
            textColor=colors.HexColor(self.template.secondary_color),
        ))

        styles.add(ParagraphStyle(
            name='ChineseHeading',
            fontSize=14,
            leading=18,
            spaceBefore=15,
            spaceAfter=10,
            textColor=colors.HexColor(self.template.primary_color),
        ))

        styles.add(ParagraphStyle(
            name='ChineseBody',
            fontSize=10,
            leading=14,
            spaceAfter=6,
        ))

        return styles

    def _create_table(self, data: List[List[str]], col_widths: Optional[List[float]] = None) -> Table:
        table = Table(data, colWidths=col_widths)

        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.template.primary_color)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dddddd')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
        ])

        table.setStyle(style)
        return table

    def generate_single_stock_report(self, data: StockReportData, output_path: str) -> bool:
        try:
            os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)

            doc = SimpleDocTemplate(
                output_path,
                pagesize=self._get_page_size(),
                topMargin=self.template.margin_top * cm,
                bottomMargin=self.template.margin_bottom * cm,
                leftMargin=self.template.margin_left * cm,
                rightMargin=self.template.margin_right * cm,
            )

            styles = self._get_styles()
            story = []

            # 标题
            story.append(Paragraph(self.template.title, styles['ChineseTitle']))
            story.append(Paragraph(f"{data.stock_code} {data.stock_name}", styles['ChineseSubtitle']))
            story.append(Spacer(1, 20))

            # 摘要
            if self.template.include_summary:
                story.append(Paragraph("【投资摘要】", styles['ChineseHeading']))
                summary_data = [
                    ["指标", "数值"],
                    ["综合评分", f"{data.overall_score:.1f}/100" if data.overall_score else "N/A"],
                    ["最终信号", data.final_signal or "N/A"],
                    ["投资决策", data.decision or "N/A"],
                    ["ML评分", f"{data.ml_score:.1f}/10" if data.ml_score else "N/A"],
                ]
                story.append(self._create_table(summary_data, [200, 200]))
                story.append(Spacer(1, 15))

            # 财务指标
            if self.template.include_financials:
                story.append(Paragraph("【财务指标】", styles['ChineseHeading']))
                fin_data = [
                    ["指标", "数值"],
                    ["当前价格", f"¥{data.current_price:.2f}" if data.current_price else "N/A"],
                    ["PE比率", f"{data.pe_ratio:.2f}" if data.pe_ratio else "N/A"],
                    ["PB比率", f"{data.pb_ratio:.2f}" if data.pb_ratio else "N/A"],
                    ["ROE", f"{data.roe:.2%}" if data.roe else "N/A"],
                    ["毛利率", f"{data.gross_margin:.2%}" if data.gross_margin else "N/A"],
                    ["负债率", f"{data.debt_ratio:.2%}" if data.debt_ratio else "N/A"],
                ]
                story.append(self._create_table(fin_data, [200, 200]))
                story.append(Spacer(1, 15))

            # 估值分析
            if self.template.include_valuation:
                story.append(Paragraph("【估值分析】", styles['ChineseHeading']))
                val_data = [
                    ["指标", "数值"],
                    ["内在价值", f"¥{data.intrinsic_value:.2f}" if data.intrinsic_value else "N/A"],
                    ["合理价格", f"¥{data.fair_price:.2f}" if data.fair_price else "N/A"],
                    ["安全边际", f"{data.margin_of_safety:.2f}%" if data.margin_of_safety else "N/A"],
                    ["估值评分", f"{data.valuation_score:.1f}/10" if data.valuation_score else "N/A"],
                ]
                story.append(self._create_table(val_data, [200, 200]))
                story.append(Spacer(1, 15))

            # 风险评估
            if self.template.include_risk:
                story.append(Paragraph("【风险评估】", styles['ChineseHeading']))
                risk_data = [
                    ["指标", "数值"],
                    ["风险等级", data.risk_level or "N/A"],
                    ["杠杆风险", f"{data.leverage_risk:.2f}" if data.leverage_risk else "N/A"],
                    ["止损价", f"¥{data.stop_loss:.2f}" if data.stop_loss else "N/A"],
                    ["止盈价", f"¥{data.take_profit:.2f}" if data.take_profit else "N/A"],
                ]
                story.append(self._create_table(risk_data, [200, 200]))

            # 页脚
            story.append(Spacer(1, 30))
            footer = f"{self.template.footer_text}"
            if self.template.show_generation_time:
                footer += f" | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            story.append(Paragraph(footer, styles['ChineseBody']))

            doc.build(story)
            logger.info(f"PDF 报告已生成: {output_path}")
            return True
        except Exception as e:
            logger.error(f"生成 PDF 报告失败: {str(e)}")
            return False

    def generate_portfolio_report(self, data: PortfolioReportData, output_path: str) -> bool:
        try:
            os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)

            doc = SimpleDocTemplate(
                output_path,
                pagesize=self._get_page_size(),
                topMargin=self.template.margin_top * cm,
                bottomMargin=self.template.margin_bottom * cm,
                leftMargin=self.template.margin_left * cm,
                rightMargin=self.template.margin_right * cm,
            )

            styles = self._get_styles()
            story = []

            # 标题
            story.append(Paragraph(self.template.title, styles['ChineseTitle']))
            story.append(Paragraph("投资组合分析报告", styles['ChineseSubtitle']))
            story.append(Spacer(1, 20))

            # 统计摘要
            story.append(Paragraph("【组合统计】", styles['ChineseHeading']))
            stats_data = [
                ["指标", "数值"],
                ["分析股票数", str(data.total_stocks)],
                ["强烈买入", str(data.strong_buy_count)],
                ["买入", str(data.buy_count)],
                ["持有", str(data.hold_count)],
                ["卖出", str(data.sell_count)],
                ["强烈卖出", str(data.strong_sell_count)],
            ]
            story.append(self._create_table(stats_data, [200, 200]))
            story.append(Spacer(1, 20))

            # 股票列表
            story.append(Paragraph("【股票明细】", styles['ChineseHeading']))
            stock_data = [["代码", "信号", "评分", "安全边际", "ML评分"]]
            for stock in data.stocks:
                stock_data.append([
                    stock.stock_code,
                    stock.final_signal or "N/A",
                    f"{stock.overall_score:.1f}" if stock.overall_score else "N/A",
                    f"{stock.margin_of_safety:.1f}%" if stock.margin_of_safety else "N/A",
                    f"{stock.ml_score:.1f}" if stock.ml_score else "N/A",
                ])
            story.append(self._create_table(stock_data, [80, 100, 80, 80, 80]))

            # 页脚
            story.append(Spacer(1, 30))
            footer = f"{self.template.footer_text}"
            if self.template.show_generation_time:
                footer += f" | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            story.append(Paragraph(footer, styles['ChineseBody']))

            doc.build(story)
            logger.info(f"PDF 组合报告已生成: {output_path}")
            return True
        except Exception as e:
            logger.error(f"生成 PDF 组合报告失败: {str(e)}")
            return False


class ExcelReportGenerator(BaseReportGenerator):
    """Excel 报告生成器"""

    def __init__(self, template: Optional[ReportTemplate] = None):
        super().__init__(template)
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl 不可用，请运行: pip install openpyxl")

    def _apply_header_style(self, cell):
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color=self.template.primary_color.replace("#", ""),
                                 end_color=self.template.primary_color.replace("#", ""),
                                 fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _apply_cell_style(self, cell, row_idx: int):
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if row_idx % 2 == 0:
            cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    def generate_single_stock_report(self, data: StockReportData, output_path: str) -> bool:
        try:
            os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "股票分析"

            # 标题
            ws.merge_cells('A1:D1')
            ws['A1'] = f"{self.template.title} - {data.stock_code}"
            ws['A1'].font = Font(bold=True, size=16, color=self.template.primary_color.replace("#", ""))
            ws['A1'].alignment = Alignment(horizontal="center")

            # 数据行
            rows = [
                ["指标", "数值", "指标", "数值"],
                ["股票代码", data.stock_code, "股票名称", data.stock_name],
                ["当前价格", data.current_price, "PE比率", data.pe_ratio],
                ["PB比率", data.pb_ratio, "ROE", f"{data.roe:.2%}" if data.roe else None],
                ["毛利率", f"{data.gross_margin:.2%}" if data.gross_margin else None, "负债率", f"{data.debt_ratio:.2%}" if data.debt_ratio else None],
                ["内在价值", data.intrinsic_value, "合理价格", data.fair_price],
                ["安全边际", f"{data.margin_of_safety:.2f}%" if data.margin_of_safety else None, "估值评分", data.valuation_score],
                ["护城河评分", data.moat_score, "风险等级", data.risk_level],
                ["综合评分", data.overall_score, "ML评分", data.ml_score],
                ["最终信号", data.final_signal, "投资决策", data.decision],
                ["建议仓位", f"{data.position_size:.2%}" if data.position_size else None, "止损价", data.stop_loss],
            ]

            for row_idx, row in enumerate(rows, start=3):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value if value is not None else "N/A")
                    if row_idx == 3:
                        self._apply_header_style(cell)
                    else:
                        self._apply_cell_style(cell, row_idx)

            # 调整列宽
            for col in range(1, 5):
                ws.column_dimensions[get_column_letter(col)].width = 18

            # 页脚
            footer_row = len(rows) + 5
            ws.cell(row=footer_row, column=1, value=f"{self.template.footer_text} | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

            wb.save(output_path)
            logger.info(f"Excel 报告已生成: {output_path}")
            return True
        except Exception as e:
            logger.error(f"生成 Excel 报告失败: {str(e)}")
            return False

    def generate_portfolio_report(self, data: PortfolioReportData, output_path: str) -> bool:
        try:
            os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)

            wb = openpyxl.Workbook()

            # Sheet 1: 统计摘要
            ws_summary = wb.active
            ws_summary.title = "组合统计"

            ws_summary.merge_cells('A1:B1')
            ws_summary['A1'] = f"{self.template.title} - 投资组合"
            ws_summary['A1'].font = Font(bold=True, size=16, color=self.template.primary_color.replace("#", ""))

            summary_rows = [
                ["指标", "数值"],
                ["分析股票数", data.total_stocks],
                ["强烈买入", data.strong_buy_count],
                ["买入", data.buy_count],
                ["持有", data.hold_count],
                ["卖出", data.sell_count],
                ["强烈卖出", data.strong_sell_count],
                ["投资策略", data.strategy],
                ["预期收益", f"{data.expected_return:.2%}" if data.expected_return else "N/A"],
                ["风险评分", f"{data.risk_score:.1f}" if data.risk_score else "N/A"],
            ]

            for row_idx, row in enumerate(summary_rows, start=3):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws_summary.cell(row=row_idx, column=col_idx, value=value if value is not None else "N/A")
                    if row_idx == 3:
                        self._apply_header_style(cell)
                    else:
                        self._apply_cell_style(cell, row_idx)

            ws_summary.column_dimensions['A'].width = 15
            ws_summary.column_dimensions['B'].width = 15

            # Sheet 2: 股票明细
            ws_stocks = wb.create_sheet("股票明细")

            headers = ["代码", "名称", "当前价", "信号", "评分", "安全边际", "ML评分", "决策"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_stocks.cell(row=1, column=col_idx, value=header)
                self._apply_header_style(cell)

            for row_idx, stock in enumerate(data.stocks, start=2):
                values = [
                    stock.stock_code,
                    stock.stock_name,
                    stock.current_price,
                    stock.final_signal,
                    stock.overall_score,
                    f"{stock.margin_of_safety:.1f}%" if stock.margin_of_safety else "N/A",
                    stock.ml_score,
                    stock.decision,
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws_stocks.cell(row=row_idx, column=col_idx, value=value if value is not None else "N/A")
                    self._apply_cell_style(cell, row_idx)

            for col in range(1, len(headers) + 1):
                ws_stocks.column_dimensions[get_column_letter(col)].width = 12

            wb.save(output_path)
            logger.info(f"Excel 组合报告已生成: {output_path}")
            return True
        except Exception as e:
            logger.error(f"生成 Excel 组合报告失败: {str(e)}")
            return False


class ReportManager:
    """报告管理器 - 统一入口"""

    def __init__(self, template: Optional[ReportTemplate] = None):
        self.template = template or ReportTemplate()
        self._pdf_generator: Optional[PDFReportGenerator] = None
        self._excel_generator: Optional[ExcelReportGenerator] = None

    @property
    def pdf_generator(self) -> Optional[PDFReportGenerator]:
        if self._pdf_generator is None and REPORTLAB_AVAILABLE:
            self._pdf_generator = PDFReportGenerator(self.template)
        return self._pdf_generator

    @property
    def excel_generator(self) -> Optional[ExcelReportGenerator]:
        if self._excel_generator is None and OPENPYXL_AVAILABLE:
            self._excel_generator = ExcelReportGenerator(self.template)
        return self._excel_generator

    def set_template(self, template: ReportTemplate) -> None:
        self.template = template
        self._pdf_generator = None
        self._excel_generator = None

    def load_template(self, path: str) -> None:
        self.template = ReportTemplate.load(path)
        self._pdf_generator = None
        self._excel_generator = None

    def generate_pdf(self, data: StockReportData, output_path: str) -> bool:
        if self.pdf_generator:
            return self.pdf_generator.generate_single_stock_report(data, output_path)
        logger.error("PDF 生成器不可用")
        return False

    def generate_excel(self, data: StockReportData, output_path: str) -> bool:
        if self.excel_generator:
            return self.excel_generator.generate_single_stock_report(data, output_path)
        logger.error("Excel 生成器不可用")
        return False

    def generate_portfolio_pdf(self, data: PortfolioReportData, output_path: str) -> bool:
        if self.pdf_generator:
            return self.pdf_generator.generate_portfolio_report(data, output_path)
        logger.error("PDF 生成器不可用")
        return False

    def generate_portfolio_excel(self, data: PortfolioReportData, output_path: str) -> bool:
        if self.excel_generator:
            return self.excel_generator.generate_portfolio_report(data, output_path)
        logger.error("Excel 生成器不可用")
        return False

    def generate_all(self, data: StockReportData, output_dir: str, base_name: str = "report") -> Dict[str, bool]:
        results = {}

        if REPORTLAB_AVAILABLE:
            pdf_path = os.path.join(output_dir, f"{base_name}.pdf")
            results["pdf"] = self.generate_pdf(data, pdf_path)

        if OPENPYXL_AVAILABLE:
            excel_path = os.path.join(output_dir, f"{base_name}.xlsx")
            results["excel"] = self.generate_excel(data, excel_path)

        return results


# 便捷函数
def create_report_manager(template_path: Optional[str] = None) -> ReportManager:
    if template_path and os.path.exists(template_path):
        template = ReportTemplate.load(template_path)
    else:
        template = ReportTemplate()
    return ReportManager(template)
